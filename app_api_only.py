#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
å­—èŠ‚è·³åŠ¨æ‹›è˜ä¿¡æ¯ç›‘æ§ç³»ç»Ÿ - çº¯APIç‰ˆæœ¬
æœ€ç®€å•å¯é çš„è§£å†³æ–¹æ¡ˆï¼Œå®Œå…¨é¿å…æµè§ˆå™¨ä¾èµ–
"""

import os
import json
import logging
import time
import random
from datetime import datetime
from flask import Flask, render_template, jsonify
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
from openpyxl import Workbook

# é…ç½®æ—¥å¿—
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('app.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# é…ç½®
DATA_DIR = os.getenv('DATA_DIR', './data')
CACHE_FILE = os.path.join(DATA_DIR, 'bytedance_jobs_cache.json')
EXCEL_FILE = os.path.join(DATA_DIR, 'bytedance_jobs_tracker.xlsx')

# ç¡®ä¿æ•°æ®ç›®å½•å­˜åœ¨
os.makedirs(DATA_DIR, exist_ok=True)

def create_session():
    """åˆ›å»ºå¸¦é‡è¯•æœºåˆ¶çš„requestsä¼šè¯"""
    session = requests.Session()
    
    # è®¾ç½®é‡è¯•ç­–ç•¥
    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
    )
    
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    
    # è®¾ç½®è¯·æ±‚å¤´
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin'
    })
    
    return session

def fetch_jobs_api():
    """ä½¿ç”¨APIæ–¹å¼è·å–æ‹›è˜ä¿¡æ¯"""
    session = create_session()
    jobs = []
    
    try:
        logger.info("å¼€å§‹ä½¿ç”¨APIè·å–æ‹›è˜ä¿¡æ¯...")
        
        # å°è¯•å¤šä¸ªå¯èƒ½çš„APIç«¯ç‚¹
        api_urls = [
            "https://jobs.bytedance.com/api/v1/web/job/list",
            "https://jobs.bytedance.com/api/web/job/list",
            "https://job.bytedance.com/api/v1/web/job/list",
            "https://careers-api.bytedance.com/api/v1/jobs"
        ]
        
        # å¸¸ç”¨çš„æŸ¥è¯¢å‚æ•°
        params_list = [
            {
                'limit': 20,
                'offset': 0,
                'keyword': '',
                'category': '',
                'location': '',
                'type': 'experienced'
            },
            {
                'page': 1,
                'size': 20,
                'job_type': 'experienced'
            },
            {
                'limit': 20,
                'page': 1
            }
        ]
        
        for api_url in api_urls:
            for params in params_list:
                try:
                    logger.info(f"å°è¯•API: {api_url}")
                    
                    # æ·»åŠ éšæœºå»¶è¿Ÿ
                    time.sleep(random.uniform(1, 3))
                    
                    response = session.get(api_url, params=params, timeout=15)
                    
                    if response.status_code == 200:
                        try:
                            data = response.json()
                            logger.info(f"APIå“åº”æˆåŠŸ: {api_url}")
                            
                            # å°è¯•è§£æä¸åŒçš„æ•°æ®ç»“æ„
                            job_data = None
                            if 'data' in data:
                                if isinstance(data['data'], list):
                                    job_data = data['data']
                                elif 'jobs' in data['data']:
                                    job_data = data['data']['jobs']
                                elif 'list' in data['data']:
                                    job_data = data['data']['list']
                            elif 'jobs' in data:
                                job_data = data['jobs']
                            elif isinstance(data, list):
                                job_data = data
                            
                            if job_data and len(job_data) > 0:
                                logger.info(f"æ‰¾åˆ° {len(job_data)} ä¸ªèŒä½")
                                
                                for item in job_data[:20]:  # é™åˆ¶20ä¸ª
                                    try:
                                        job = {
                                            'title': item.get('title', item.get('job_title', item.get('name', 'æœªçŸ¥èŒä½'))),
                                            'location': item.get('location', item.get('city', item.get('work_location', 'æœªçŸ¥åœ°ç‚¹'))),
                                            'department': item.get('department', item.get('team', item.get('category', 'æœªçŸ¥éƒ¨é—¨'))),
                                            'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                            'source': 'api_direct',
                                            'job_id': item.get('id', item.get('job_id', '')),
                                            'description': item.get('description', item.get('requirement', ''))[:200]  # é™åˆ¶é•¿åº¦
                                        }
                                        
                                        if job['title'] and job['title'] != 'æœªçŸ¥èŒä½':
                                            jobs.append(job)
                                    except Exception as e:
                                        logger.debug(f"è§£æå•ä¸ªèŒä½å¤±è´¥: {e}")
                                        continue
                                
                                if jobs:
                                    logger.info(f"APIæˆåŠŸè·å– {len(jobs)} ä¸ªèŒä½")
                                    return jobs
                            
                        except json.JSONDecodeError as e:
                            logger.debug(f"JSONè§£æå¤±è´¥: {e}")
                            continue
                    
                    elif response.status_code == 403:
                        logger.warning(f"APIè®¿é—®è¢«æ‹’ç» (403): {api_url}")
                    elif response.status_code == 404:
                        logger.debug(f"APIä¸å­˜åœ¨ (404): {api_url}")
                    else:
                        logger.debug(f"APIå“åº”å¼‚å¸¸ ({response.status_code}): {api_url}")
                        
                except requests.exceptions.RequestException as e:
                    logger.debug(f"è¯·æ±‚å¤±è´¥: {api_url} - {e}")
                    continue
                except Exception as e:
                    logger.debug(f"å¤„ç†APIå¤±è´¥: {api_url} - {e}")
                    continue
        
        # å¦‚æœæ‰€æœ‰APIéƒ½å¤±è´¥ï¼Œè¿”å›æ¨¡æ‹Ÿæ•°æ®
        if not jobs:
            logger.warning("æ‰€æœ‰APIå°è¯•å¤±è´¥ï¼Œè¿”å›æ¨¡æ‹Ÿæ•°æ®")
            jobs = generate_mock_jobs()
        
        return jobs
        
    except Exception as e:
        logger.error(f"APIè·å–è¿‡ç¨‹å¤±è´¥: {e}")
        return generate_mock_jobs()
    finally:
        session.close()

def generate_mock_jobs():
    """ç”Ÿæˆæ¨¡æ‹ŸèŒä½æ•°æ®"""
    mock_jobs = [
        {
            'title': 'å‰ç«¯å¼€å‘å·¥ç¨‹å¸ˆ',
            'location': 'åŒ—äº¬',
            'department': 'æŠ€æœ¯éƒ¨',
            'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'source': 'mock_data',
            'job_id': 'mock_001',
            'description': 'è´Ÿè´£å‰ç«¯é¡µé¢å¼€å‘å’Œç”¨æˆ·ä½“éªŒä¼˜åŒ–'
        },
        {
            'title': 'åç«¯å¼€å‘å·¥ç¨‹å¸ˆ',
            'location': 'ä¸Šæµ·',
            'department': 'æŠ€æœ¯éƒ¨',
            'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'source': 'mock_data',
            'job_id': 'mock_002',
            'description': 'è´Ÿè´£åç«¯æœåŠ¡å¼€å‘å’Œç³»ç»Ÿæ¶æ„è®¾è®¡'
        },
        {
            'title': 'äº§å“ç»ç†',
            'location': 'æ·±åœ³',
            'department': 'äº§å“éƒ¨',
            'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'source': 'mock_data',
            'job_id': 'mock_003',
            'description': 'è´Ÿè´£äº§å“è§„åˆ’å’Œéœ€æ±‚åˆ†æ'
        },
        {
            'title': 'æ•°æ®åˆ†æå¸ˆ',
            'location': 'æ­å·',
            'department': 'æ•°æ®éƒ¨',
            'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'source': 'mock_data',
            'job_id': 'mock_004',
            'description': 'è´Ÿè´£æ•°æ®æŒ–æ˜å’Œä¸šåŠ¡åˆ†æ'
        },
        {
            'title': 'UIè®¾è®¡å¸ˆ',
            'location': 'å¹¿å·',
            'department': 'è®¾è®¡éƒ¨',
            'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'source': 'mock_data',
            'job_id': 'mock_005',
            'description': 'è´Ÿè´£ç•Œé¢è®¾è®¡å’Œç”¨æˆ·ä½“éªŒè®¾è®¡'
        }
    ]
    
    logger.info(f"ç”Ÿæˆ {len(mock_jobs)} ä¸ªæ¨¡æ‹ŸèŒä½æ•°æ®")
    return mock_jobs

def save_to_cache(jobs):
    """ä¿å­˜åˆ°ç¼“å­˜æ–‡ä»¶"""
    try:
        cache_data = {
            'jobs': jobs,
            'last_update': datetime.now().isoformat(),
            'total_count': len(jobs),
            'version': 'api_only'
        }
        
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"ç¼“å­˜å·²æ›´æ–°ï¼Œå…± {len(jobs)} ä¸ªèŒä½")
    except Exception as e:
        logger.error(f"ä¿å­˜ç¼“å­˜å¤±è´¥: {e}")

def load_from_cache():
    """ä»ç¼“å­˜åŠ è½½æ•°æ®"""
    try:
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"åŠ è½½ç¼“å­˜å¤±è´¥: {e}")
    return None

def save_to_excel(jobs):
    """ä¿å­˜åˆ°Excelæ–‡ä»¶"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "å­—èŠ‚è·³åŠ¨æ‹›è˜ä¿¡æ¯"
            # æ·»åŠ è¡¨å¤´
            headers = ['èŒä½åç§°', 'å·¥ä½œåœ°ç‚¹', 'éƒ¨é—¨', 'æ›´æ–°æ—¶é—´', 'æ•°æ®æº', 'èŒä½ID', 'èŒä½æè¿°']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
        
        # æ¸…é™¤æ—§æ•°æ®ï¼ˆä¿ç•™è¡¨å¤´ï¼‰
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        
        # æ·»åŠ æ–°æ•°æ®
        for row_idx, job in enumerate(jobs, 2):
            ws.cell(row=row_idx, column=1, value=job.get('title', ''))
            ws.cell(row=row_idx, column=2, value=job.get('location', ''))
            ws.cell(row=row_idx, column=3, value=job.get('department', ''))
            ws.cell(row=row_idx, column=4, value=job.get('update_time', ''))
            ws.cell(row=row_idx, column=5, value=job.get('source', ''))
            ws.cell(row=row_idx, column=6, value=job.get('job_id', ''))
            ws.cell(row=row_idx, column=7, value=job.get('description', ''))
        
        wb.save(EXCEL_FILE)
        logger.info(f"Excelæ–‡ä»¶å·²æ›´æ–°ï¼Œå…± {len(jobs)} ä¸ªèŒä½")
    except Exception as e:
        logger.error(f"ä¿å­˜Excelå¤±è´¥: {e}")

@app.route('/')
def index():
    """ä¸»é¡µ"""
    return render_template('index.html')

@app.route('/api/jobs')
def get_jobs():
    """è·å–èŒä½ä¿¡æ¯API"""
    try:
        # å°è¯•ä»ç¼“å­˜åŠ è½½
        cache_data = load_from_cache()
        
        # å¦‚æœç¼“å­˜ä¸å­˜åœ¨æˆ–è¶…è¿‡30åˆ†é’Ÿï¼Œé‡æ–°è·å–
        should_refresh = True
        if cache_data and 'last_update' in cache_data:
            try:
                last_update = datetime.fromisoformat(cache_data['last_update'])
                if (datetime.now() - last_update).seconds < 1800:  # 30åˆ†é’Ÿå†…
                    should_refresh = False
            except:
                pass
        
        if should_refresh:
            logger.info("ç¼“å­˜è¿‡æœŸæˆ–ä¸å­˜åœ¨ï¼Œé‡æ–°è·å–æ•°æ®...")
            jobs = fetch_jobs_api()
            if jobs:
                save_to_cache(jobs)
                save_to_excel(jobs)
                cache_data = {
                    'jobs': jobs,
                    'last_update': datetime.now().isoformat(),
                    'total_count': len(jobs),
                    'version': 'api_only'
                }
            else:
                # å¦‚æœè·å–å¤±è´¥ï¼Œä½¿ç”¨ç¼“å­˜æ•°æ®æˆ–ç”Ÿæˆæ¨¡æ‹Ÿæ•°æ®
                if not cache_data:
                    jobs = generate_mock_jobs()
                    cache_data = {
                        'jobs': jobs,
                        'last_update': datetime.now().isoformat(),
                        'total_count': len(jobs),
                        'version': 'api_only',
                        'error': 'APIè·å–å¤±è´¥ï¼Œä½¿ç”¨æ¨¡æ‹Ÿæ•°æ®'
                    }
        
        return jsonify({
            'success': True,
            'data': cache_data,
            'message': 'APIçº¯å‡€ç‰ˆæœ¬è¿è¡Œä¸­ï¼Œæ— æµè§ˆå™¨ä¾èµ–'
        })
        
    except Exception as e:
        logger.error(f"APIé”™è¯¯: {e}")
        # è¿”å›æ¨¡æ‹Ÿæ•°æ®ç¡®ä¿æœåŠ¡å¯ç”¨
        mock_jobs = generate_mock_jobs()
        return jsonify({
            'success': True,
            'data': {
                'jobs': mock_jobs,
                'last_update': datetime.now().isoformat(),
                'total_count': len(mock_jobs),
                'version': 'api_only_fallback'
            },
            'message': 'æœåŠ¡è¿è¡Œæ­£å¸¸ï¼Œå½“å‰æ˜¾ç¤ºæ¨¡æ‹Ÿæ•°æ®'
        })

@app.route('/api/refresh')
def refresh_jobs():
    """å¼ºåˆ¶åˆ·æ–°èŒä½ä¿¡æ¯"""
    try:
        logger.info("æ‰‹åŠ¨åˆ·æ–°èŒä½ä¿¡æ¯...")
        jobs = fetch_jobs_api()
        
        if jobs:
            save_to_cache(jobs)
            save_to_excel(jobs)
            return jsonify({
                'success': True,
                'message': f'æˆåŠŸåˆ·æ–° {len(jobs)} ä¸ªèŒä½ä¿¡æ¯',
                'count': len(jobs),
                'source': jobs[0].get('source', 'unknown') if jobs else 'none'
            })
        else:
            # å³ä½¿è·å–å¤±è´¥ï¼Œä¹Ÿè¿”å›æ¨¡æ‹Ÿæ•°æ®ç¡®ä¿æœåŠ¡å¯ç”¨
            mock_jobs = generate_mock_jobs()
            save_to_cache(mock_jobs)
            return jsonify({
                'success': True,
                'message': f'APIè·å–å¤±è´¥ï¼Œè¿”å› {len(mock_jobs)} ä¸ªæ¨¡æ‹ŸèŒä½',
                'count': len(mock_jobs),
                'source': 'mock_data'
            })
            
    except Exception as e:
        logger.error(f"åˆ·æ–°å¤±è´¥: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'åˆ·æ–°è¿‡ç¨‹é‡åˆ°é”™è¯¯'
        }), 500

@app.route('/health')
def health_check():
    """å¥åº·æ£€æŸ¥"""
    try:
        # æµ‹è¯•ç½‘ç»œè¿æ¥
        session = create_session()
        try:
            response = session.get('https://www.baidu.com', timeout=5)
            network_status = "æ­£å¸¸" if response.status_code == 200 else "å¼‚å¸¸"
        except:
            network_status = "å¼‚å¸¸"
        finally:
            session.close()
        
        return jsonify({
            'status': 'healthy',
            'network': network_status,
            'version': 'api_only_v1.0',
            'features': ['çº¯APIè·å–', 'æ— æµè§ˆå™¨ä¾èµ–', 'æ¨¡æ‹Ÿæ•°æ®å¤‡ä»½'],
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

if __name__ == '__main__':
    logger.info("=" * 50)
    logger.info("ğŸš€ å­—èŠ‚è·³åŠ¨æ‹›è˜ç›‘æ§ç³»ç»Ÿ - çº¯APIç‰ˆæœ¬ v2.0")
    logger.info("âœ… æ— æµè§ˆå™¨ä¾èµ–ï¼Œè½»é‡çº§éƒ¨ç½²")
    logger.info("âœ… æ”¯æŒæ¨¡æ‹Ÿæ•°æ®å¤‡ä»½ï¼Œç¡®ä¿æœåŠ¡å¯ç”¨")
    logger.info("âœ… å®Œå…¨é¿å…Playwright/Seleniumé—®é¢˜")
    logger.info("=" * 50)

    port = int(os.getenv('PORT', 8080))
    logger.info(f"ğŸŒ æœåŠ¡å¯åŠ¨åœ¨ç«¯å£: {port}")
    app.run(host='0.0.0.0', port=port, debug=False)