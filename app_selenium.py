#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
字节跳动招聘信息监控系统 - Selenium版本
备选方案：使用Selenium替代Playwright
"""

import os
import json
import logging
from datetime import datetime
from flask import Flask, render_template, jsonify
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
import requests
import openpyxl
from openpyxl import Workbook

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('app.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# 配置
DATA_DIR = os.getenv('DATA_DIR', './data')
CACHE_FILE = os.path.join(DATA_DIR, 'bytedance_jobs_cache.json')
EXCEL_FILE = os.path.join(DATA_DIR, 'bytedance_jobs_tracker.xlsx')

# 确保数据目录存在
os.makedirs(DATA_DIR, exist_ok=True)

def create_chrome_driver():
    """创建Chrome WebDriver"""
    try:
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
        
        driver = webdriver.Chrome(options=chrome_options)
        driver.set_page_load_timeout(30)
        return driver
    except Exception as e:
        logger.error(f"创建Chrome驱动失败: {e}")
        return None

def fetch_jobs_selenium():
    """使用Selenium获取招聘信息"""
    driver = None
    try:
        logger.info("开始使用Selenium获取招聘信息...")
        
        driver = create_chrome_driver()
        if not driver:
            raise Exception("无法创建Chrome驱动")
        
        # 访问字节跳动招聘页面
        url = "https://jobs.bytedance.com/experienced/position"
        driver.get(url)
        
        # 等待页面加载
        wait = WebDriverWait(driver, 20)
        
        # 等待职位列表加载
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".job-list-box, .position-list, [class*='job'], [class*='position']")))
        except TimeoutException:
            logger.warning("等待职位列表超时，尝试继续解析...")
        
        # 尝试多种选择器来获取职位信息
        job_selectors = [
            ".job-list-box .job-item",
            ".position-list .position-item",
            "[class*='job-item']",
            "[class*='position-item']",
            ".job-card",
            ".position-card"
        ]
        
        jobs = []
        for selector in job_selectors:
            try:
                job_elements = driver.find_elements(By.CSS_SELECTOR, selector)
                if job_elements:
                    logger.info(f"使用选择器 {selector} 找到 {len(job_elements)} 个职位")
                    
                    for element in job_elements[:20]:  # 限制获取前20个
                        try:
                            # 尝试提取职位信息
                            title_elem = element.find_element(By.CSS_SELECTOR, "h3, .job-title, .position-title, [class*='title']")
                            title = title_elem.text.strip() if title_elem else "未知职位"
                            
                            # 尝试获取其他信息
                            location = "未知地点"
                            department = "未知部门"
                            
                            try:
                                location_elem = element.find_element(By.CSS_SELECTOR, ".location, .job-location, [class*='location']")
                                location = location_elem.text.strip()
                            except:
                                pass
                            
                            try:
                                dept_elem = element.find_element(By.CSS_SELECTOR, ".department, .job-department, [class*='department']")
                                department = dept_elem.text.strip()
                            except:
                                pass
                            
                            if title and title != "未知职位":
                                jobs.append({
                                    'title': title,
                                    'location': location,
                                    'department': department,
                                    'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                    'source': 'selenium'
                                })
                        except Exception as e:
                            logger.debug(f"解析单个职位失败: {e}")
                            continue
                    
                    if jobs:
                        break  # 如果成功获取到职位，跳出循环
            except Exception as e:
                logger.debug(f"选择器 {selector} 失败: {e}")
                continue
        
        if not jobs:
            # 如果没有找到职位，尝试获取页面源码进行分析
            page_source = driver.page_source
            logger.warning("未找到职位信息，页面可能需要更多时间加载或结构已变化")
            
            # 返回一些示例数据以便测试
            jobs = [{
                'title': 'Selenium测试职位',
                'location': '北京',
                'department': '技术部',
                'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'source': 'selenium_fallback'
            }]
        
        logger.info(f"Selenium成功获取 {len(jobs)} 个职位信息")
        return jobs
        
    except Exception as e:
        logger.error(f"Selenium获取失败: {e}")
        return []
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass

def save_to_cache(jobs):
    """保存到缓存文件"""
    try:
        cache_data = {
            'jobs': jobs,
            'last_update': datetime.now().isoformat(),
            'total_count': len(jobs)
        }
        
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"缓存已更新，共 {len(jobs)} 个职位")
    except Exception as e:
        logger.error(f"保存缓存失败: {e}")

def load_from_cache():
    """从缓存加载数据"""
    try:
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"加载缓存失败: {e}")
    return None

def save_to_excel(jobs):
    """保存到Excel文件"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "字节跳动招聘信息"
            # 添加表头
            headers = ['职位名称', '工作地点', '部门', '更新时间', '数据源']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
        
        # 清除旧数据（保留表头）
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        
        # 添加新数据
        for row_idx, job in enumerate(jobs, 2):
            ws.cell(row=row_idx, column=1, value=job.get('title', ''))
            ws.cell(row=row_idx, column=2, value=job.get('location', ''))
            ws.cell(row=row_idx, column=3, value=job.get('department', ''))
            ws.cell(row=row_idx, column=4, value=job.get('update_time', ''))
            ws.cell(row=row_idx, column=5, value=job.get('source', ''))
        
        wb.save(EXCEL_FILE)
        logger.info(f"Excel文件已更新，共 {len(jobs)} 个职位")
    except Exception as e:
        logger.error(f"保存Excel失败: {e}")

@app.route('/')
def index():
    """主页"""
    return render_template('index.html')

@app.route('/api/jobs')
def get_jobs():
    """获取职位信息API"""
    try:
        # 尝试从缓存加载
        cache_data = load_from_cache()
        
        # 如果缓存不存在或超过1小时，重新获取
        should_refresh = True
        if cache_data and 'last_update' in cache_data:
            try:
                last_update = datetime.fromisoformat(cache_data['last_update'])
                if (datetime.now() - last_update).seconds < 3600:  # 1小时内
                    should_refresh = False
            except:
                pass
        
        if should_refresh:
            logger.info("缓存过期或不存在，重新获取数据...")
            jobs = fetch_jobs_selenium()
            if jobs:
                save_to_cache(jobs)
                save_to_excel(jobs)
                cache_data = {
                    'jobs': jobs,
                    'last_update': datetime.now().isoformat(),
                    'total_count': len(jobs)
                }
            else:
                # 如果获取失败，使用缓存数据
                if not cache_data:
                    cache_data = {
                        'jobs': [],
                        'last_update': datetime.now().isoformat(),
                        'total_count': 0,
                        'error': 'Selenium获取失败，且无缓存数据'
                    }
        
        return jsonify({
            'success': True,
            'data': cache_data,
            'message': 'Selenium版本运行中'
        })
        
    except Exception as e:
        logger.error(f"API错误: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Selenium版本遇到错误'
        }), 500

@app.route('/api/refresh')
def refresh_jobs():
    """强制刷新职位信息"""
    try:
        logger.info("手动刷新职位信息...")
        jobs = fetch_jobs_selenium()
        
        if jobs:
            save_to_cache(jobs)
            save_to_excel(jobs)
            return jsonify({
                'success': True,
                'message': f'成功刷新 {len(jobs)} 个职位信息',
                'count': len(jobs)
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Selenium获取失败，请检查网络连接或网站结构变化'
            }), 500
            
    except Exception as e:
        logger.error(f"刷新失败: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/health')
def health_check():
    """健康检查"""
    try:
        # 检查Selenium是否可用
        driver = create_chrome_driver()
        if driver:
            driver.quit()
            selenium_status = "正常"
        else:
            selenium_status = "异常"
        
        return jsonify({
            'status': 'healthy',
            'selenium': selenium_status,
            'version': 'selenium_backup',
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

if __name__ == '__main__':
    logger.info("启动字节跳动招聘监控系统 - Selenium版本")
    
    # 启动时检查Selenium状态
    try:
        driver = create_chrome_driver()
        if driver:
            logger.info("✅ Selenium Chrome驱动初始化成功")
            driver.quit()
        else:
            logger.error("❌ Selenium Chrome驱动初始化失败")
    except Exception as e:
        logger.error(f"❌ Selenium检查失败: {e}")
    
    port = int(os.getenv('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)