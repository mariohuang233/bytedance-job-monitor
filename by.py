# bytedance_job_monitor.py

import asyncio
import hashlib
import json
import logging
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Any

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from playwright.async_api import async_playwright, Browser

# --- 1. 配置区 ---

# 日志配置# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)

# 文件名及路径配置
# 文件将保存在当前用户的"文稿" (Documents) 文件夹下
DOCUMENTS_PATH = Path.home() / "Documents"
OUTPUT_FILENAME = DOCUMENTS_PATH / "bytedance_jobs_tracker.xlsx"
JSON_CACHE_FILENAME = DOCUMENTS_PATH / "bytedance_jobs_cache.json"

# 任务配置
TASK_CONFIGS: List[Dict[str, Any]] = [
    {
        'id': 1,
        'name': '实习招聘',
        'sheet_name': 'intern',
        'url': "https://jobs.bytedance.com/campus/position?keywords=&category=6704215864629004552%2C6704215864591255820%2C6704216224387041544%2C6704215924712409352&location=CT_125&project=7481474995534301447%2C7468181472685164808%2C7194661644654577981%2C7194661126919358757&type=&job_hot_flag=&current=1&limit=2000&functionCategory=&tag=",
        'api_url_mark': "api/v1/search/job/posts",
        'extra_fields': []
    },
    {
        'id': 2,
        'name': '校园招聘',
        'sheet_name': 'campus',
        'url': "https://jobs.bytedance.com/campus/position?keywords=&category=6704215864629004552%2C6704215864591255820%2C6704216224387041544%2C6704215924712409352&location=CT_125&project=7525009396952582407&type=&job_hot_flag=&current=1&limit=2000&functionCategory=&tag=",
        'api_url_mark': "api/v1/search/job/posts",
        'extra_fields': ['location', 'department']
    },
    {
        'id': 3,
        'name': '社会招聘',
        'sheet_name': 'experienced',
        'url': "https://jobs.bytedance.com/experienced/position?keywords=&category=6704215864629004552%2C6704215864591255820%2C6704215924712409352%2C6704216224387041544&location=CT_125&project=&type=&job_hot_flag=&current=1&limit=600&functionCategory=&tag=",
        'api_url_mark': "api/v1/search/job/posts",
        'extra_fields': ['location', 'department']
    }
]

# --- 2. 核心逻辑区 ---

class JobMonitor:
    """字节跳动职位监控器（异步版），封装了抓取、数据处理、保存和通知的全部逻辑。"""

    def __init__(self, tasks: List[Dict[str, Any]], filename: Path, headless: bool = True):
        self.tasks = tasks
        self.filename = filename
        self.json_cache_filename = JSON_CACHE_FILENAME
        self.headless = headless
        self.results: List[tuple[str, str, List[Dict[str, Any]]]] = []

    @staticmethod
    def _generate_job_hash(job_data: Dict[str, Any]) -> str:
        """为职位数据生成唯一的MD5哈希值，用于识别重复职位。"""
        key_fields = ['code', 'title', 'description', 'requirement']
        hash_string = ''.join(str(job_data.get(field, '')) for field in key_fields)
        return hashlib.md5(hash_string.encode('utf-8')).hexdigest()
    
    def _save_json_cache(self, data_frames: Dict[str, pd.DataFrame]) -> None:
        """将数据保存为JSON缓存文件。"""
        try:
            cache_data = {}
            for sheet_name, df in data_frames.items():
                # 将DataFrame转换为字典列表，处理时间格式
                records = df.to_dict('records')
                for record in records:
                    # 确保所有值都是JSON可序列化的
                    for key, value in record.items():
                        if pd.isna(value):
                            record[key] = None
                        elif isinstance(value, (pd.Timestamp, datetime)):
                            record[key] = str(value)
                cache_data[sheet_name] = records
            
            with open(self.json_cache_filename, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, ensure_ascii=False, indent=2)
            
            logging.info(f"💾 JSON缓存已保存至: {self.json_cache_filename}")
        except Exception as e:
            logging.error(f"⚠️ 保存JSON缓存时出错: {e}")
    
    def _load_json_cache(self) -> Dict[str, pd.DataFrame]:
        """从JSON缓存文件加载数据。"""
        cache_dataframes: Dict[str, pd.DataFrame] = {}
        
        if not self.json_cache_filename.exists():
            logging.info(f"JSON缓存文件 {self.json_cache_filename} 不存在。")
            return cache_dataframes
        
        try:
            with open(self.json_cache_filename, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
            
            for sheet_name, records in cache_data.items():
                if records:
                    df = pd.DataFrame(records)
                    # 清空旧的高亮标记
                    if 'highlight_time' in df.columns:
                        df['highlight_time'] = None
                    cache_dataframes[sheet_name] = df
                    logging.info(f"已从JSON缓存加载工作表 '{sheet_name}' 的 {len(records)} 条记录。")
            
        except Exception as e:
            logging.warning(f"读取JSON缓存文件 {self.json_cache_filename} 出错: {e}。将忽略缓存。")
        
        return cache_dataframes
    
    @staticmethod
    def _sort_jobs_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        """统一的岗位排序方法：新岗位(is_new=True)排在最前，然后按发布时间降序排列。"""
        if 'is_new' in df.columns:
            # 先按is_new降序(True在前)，再按publish_time降序
            df = df.sort_values(by=['is_new', 'publish_time'], ascending=[False, False])
        else:
            # 如果没有is_new字段，只按发布时间排序
            df = df.sort_values(by='publish_time', ascending=False)
        return df

    def _load_existing_hashes(self) -> tuple[Dict[str, Set[str]], Dict[str, pd.DataFrame]]:
        """从JSON缓存或Excel文件中加载每个工作表的职位哈希值和数据框。"""
        existing_hashes: Dict[str, Set[str]] = {}
        existing_dataframes: Dict[str, pd.DataFrame] = {}
        
        # 优先从JSON缓存加载
        cache_dataframes = self._load_json_cache()
        if cache_dataframes:
            logging.info("使用JSON缓存数据。")
            existing_dataframes = cache_dataframes
        elif self.filename.exists():
            logging.info("JSON缓存不存在，从Excel文件加载数据。")
            try:
                with pd.ExcelFile(self.filename, engine='openpyxl') as xls:
                    for sheet_name in xls.sheet_names:
                        df = pd.read_excel(xls, sheet_name=sheet_name)
                        
                        # 去除旧的高亮标记：清空highlight_time字段
                        if 'highlight_time' in df.columns:
                            df['highlight_time'] = None
                        
                        existing_dataframes[sheet_name] = df
                        logging.info(f"已从Excel工作表 '{sheet_name}' 加载 {len(df)} 条记录。")
            except Exception as e:
                logging.warning(f"读取Excel文件 {self.filename} 出错: {e}。将作为首次运行处理。")
        else:
            logging.info(f"Excel文件 {self.filename} 和JSON缓存都不存在，将创建新文件。")
        
        # 为所有数据框生成哈希值
        for sheet_name, df in existing_dataframes.items():
            hashes = {self._generate_job_hash(row.to_dict()) for _, row in df.iterrows()}
            existing_hashes[sheet_name] = hashes
            logging.info(f"已为工作表 '{sheet_name}' 生成 {len(hashes)} 个职位哈希。")
        
        return existing_hashes, existing_dataframes

    async def _run_single_task_async(self, task_config: Dict[str, Any], browser: Browser) -> None:
        """在独立的浏览器上下文中异步运行单个抓取任务。"""
        task_name = task_config['name']
        sheet_name = task_config['sheet_name']
        scraped_jobs: List[Dict[str, Any]] = []
        context = None
        
        try:
            context = await browser.new_context(
                user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"
            )
            page = await context.new_page()
            logging.info(f"🚀 开始任务: {task_name}")

            async with page.expect_response(lambda r: task_config['api_url_mark'] in r.url, timeout=30000) as response_info:
                await page.goto(task_config['url'], wait_until="domcontentloaded")
            
            response = await response_info.value
            if response.status == 200:
                data = await response.json()
                job_list = data.get("data", {}).get("job_post_list", [])
                
                # 调试：打印第一个职位的完整数据结构
                if job_list and logging.getLogger().isEnabledFor(logging.DEBUG):
                    logging.debug(f"API返回的第一个职位完整数据: {json.dumps(job_list[0], ensure_ascii=False, indent=2)}")
                
                for job in job_list:
                    publish_time = datetime.fromtimestamp(job["publish_time"] / 1000)
                    
                    # 扩展job_info，包含更多API字段
                    job_info = {
                        # 基础信息
                        "title": job.get("title"),
                        "sub_title": job.get("sub_title"),
                        "description": job.get("description"),
                        "requirement": job.get("requirement"),
                        "publish_time": publish_time.strftime("%Y-%m-%d %H:%M:%S"),
                        "code": job.get("code"),
                        
                        # 职位基本信息
                        "job_id": job.get("id"),
                        "job_type": job.get("job_type"),
                        "job_category": job.get("job_category", {}).get("name") if isinstance(job.get("job_category"), dict) else job.get("job_category"),
                        "job_function": job.get("job_function", {}).get("name") if isinstance(job.get("job_function"), dict) else job.get("job_function"),
                        "department_id": job.get("department_id"),
                        "job_process_id": job.get("job_process_id"),
                        
                        # 招聘类型和项目信息
                        "recruit_type_name": job.get("recruit_type", {}).get("name") if isinstance(job.get("recruit_type"), dict) else None,
                        "recruit_type_parent": job.get("recruit_type", {}).get("parent", {}).get("name") if isinstance(job.get("recruit_type"), dict) and job.get("recruit_type", {}).get("parent") else None,
                        "job_subject_name": job.get("job_subject", {}).get("name", {}).get("zh_cn") if isinstance(job.get("job_subject"), dict) and isinstance(job.get("job_subject", {}).get("name"), dict) else job.get("job_subject", {}).get("name") if isinstance(job.get("job_subject"), dict) else None,
                        
                        # 地理位置信息
                        "city_list": ", ".join([city.get("name", "") for city in job.get("city_list", []) if isinstance(city, dict)]) if job.get("city_list") else None,
                        "city_codes": ", ".join([city.get("code", "") for city in job.get("city_list", []) if isinstance(city, dict)]) if job.get("city_list") else None,
                        "address": job.get("address"),
                        
                        # 职位要求
                        "degree": job.get("degree"),
                        "experience": job.get("experience"),
                        "min_salary": job.get("min_salary"),
                        "max_salary": job.get("max_salary"),
                        "currency": job.get("currency"),
                        "head_count": job.get("head_count"),
                        
                        # 职位状态和标识
                        "job_hot_flag": job.get("job_hot_flag"),
                        "is_urgent": job.get("is_urgent"),
                        "job_active_status": job.get("job_active_status"),
                        "recommend_id": job.get("recommend_id"),
                        
                        # 其他信息
                        "team_name": job.get("team_name"),
                        "brand_name": job.get("brand_name"),
                        "ats_online_apply": job.get("ats_online_apply"),
                        "pc_job_url": job.get("pc_job_url"),
                        "wap_job_url": job.get("wap_job_url"),
                        "storefront_mode": job.get("storefront_mode"),
                        "process_type": job.get("process_type"),
                    }
                    
                    # 处理配置中的额外字段
                    for field in task_config['extra_fields']:
                        value = job.get(field)
                        job_info[field] = value.get('name') if isinstance(value, dict) else value
                    
                    # 清理None值，保持数据整洁
                    job_info = {k: v for k, v in job_info.items() if v is not None and v != ''}
                    
                    scraped_jobs.append(job_info)
                
                logging.info(f"✅ 任务 '{task_name}' 成功获取 {len(scraped_jobs)} 个职位。")
            else:
                logging.error(f"❌ 任务 '{task_name}' API 响应状态码: {response.status}")

        except Exception as e:
            logging.error(f"❌ 任务 '{task_name}' 执行失败: {e}", exc_info=False)
        finally:
            if context:
                await context.close()
            self.results.append((sheet_name, task_name, scraped_jobs))

    def _process_results(self, existing_hashes: Dict[str, Set[str]], existing_dataframes: Dict[str, pd.DataFrame]) -> Dict:
        """处理所有任务结果，合并数据并识别新职位。"""
        final_data_frames: Dict[str, pd.DataFrame] = {}
        summary_info: List[Dict[str, Any]] = []
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for sheet_name, task_name, new_jobs_data in self.results:
            previous_hashes = existing_hashes.get(sheet_name, set())
            existing_df = existing_dataframes.get(sheet_name, pd.DataFrame())
            
            if not new_jobs_data:
                # 如果没有新数据，但有旧数据，则保留旧数据
                if not existing_df.empty:
                    final_data_frames[sheet_name] = self._sort_jobs_dataframe(existing_df)
                summary_info.append({'task_name': task_name, 'new_count': 0, 'total_count': len(existing_df)})
                continue
            
            # 处理新抓取的数据
            for job in new_jobs_data:
                job_hash = self._generate_job_hash(job)
                is_new = job_hash not in previous_hashes
                job['is_new'] = is_new
                # 为新岗位添加高亮时间标记
                job['highlight_time'] = current_time if is_new else None
            
            new_df = pd.DataFrame(new_jobs_data)
            
            # 合并新旧数据
            if not existing_df.empty:
                # 确保旧数据有必要的字段
                if 'is_new' not in existing_df.columns:
                    existing_df['is_new'] = False
                if 'highlight_time' not in existing_df.columns:
                    existing_df['highlight_time'] = None
                
                # 合并数据并去重（基于哈希值）
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                # 根据职位哈希去重，保留最新的记录
                combined_df['job_hash'] = combined_df.apply(lambda row: self._generate_job_hash(row.to_dict()), axis=1)
                combined_df = combined_df.drop_duplicates(subset=['job_hash'], keep='last')
                combined_df = combined_df.drop(columns=['job_hash'])
                final_df = combined_df
            else:
                final_df = new_df
            
            final_df = self._sort_jobs_dataframe(final_df)
            final_data_frames[sheet_name] = final_df
            
            new_count = final_df['is_new'].sum() if 'is_new' in final_df.columns else 0
            summary_info.append({
                'task_name': task_name,
                'new_count': new_count,
                'total_count': len(final_df)
            })
        
        return {"data_frames": final_data_frames, "summary": summary_info}

    @staticmethod
    def _extract_team_intro(description: str) -> str:
        """从description中提取团队介绍部分。"""
        if not description:
            return ""
        
        # 查找团队介绍的结束位置
        text = description.strip()
        
        import re
        
        # 如果直接以数字编号开始，说明没有团队介绍
        if re.match(r'^[1-9]、', text):
            return "暂无团队介绍"
        
        # 寻找团队介绍的结束标志
        patterns = [
            r'团队介绍[：:][^\n]*\n\n',  # 团队介绍：...\n\n
            r'团队介绍[：:][^1-9]*(?=[1-9]、)',  # 团队介绍：...直到数字编号开始
            r'^[^1-9]*?(?=[1-9]、)',  # 从开头到第一个数字编号
            r'^.*?(?=\n\n[1-9])',  # 从开头到双换行+数字
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.MULTILINE | re.DOTALL)
            if match:
                result = match.group(0).strip()
                return result if result else "暂无团队介绍"
        
        # 如果没有找到明确的分割点，取前面的部分
        lines = text.split('\n')
        if len(lines) > 3:
            # 寻找第一个以数字开头的行
            for i, line in enumerate(lines):
                if re.match(r'^[1-9]、', line.strip()):
                    intro = '\n'.join(lines[:i]).strip()
                    return intro if intro else "暂无团队介绍"
        
        # 默认返回前200个字符
        if len(text) > 200:
            return text[:200] + '...'
        else:
            return text if text else "暂无团队介绍"
    
    @staticmethod
    def _extract_job_details(description: str) -> str:
        """从description中提取岗位细节部分。"""
        if not description:
            return ""
        
        text = description.strip()
        
        # 寻找岗位细节的开始位置
        import re
        patterns = [
            r'([1-9]、.*)',  # 从第一个数字编号开始
            r'(职责[：:].*)',  # 从职责开始
            r'(工作内容[：:].*)',  # 从工作内容开始
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.MULTILINE | re.DOTALL)
            if match:
                return match.group(1).strip()
        
        # 如果没有找到明确的开始点，尝试从团队介绍后开始
        team_intro = JobMonitor._extract_team_intro(description)
        if team_intro and len(team_intro) < len(text):
            remaining = text[len(team_intro):].strip()
            return remaining if remaining else text
        
        # 默认返回后面的部分
        return text[200:] if len(text) > 200 else text

    def _save_and_highlight(self, data_frames: Dict[str, pd.DataFrame]) -> None:
        """将数据保存到Excel和JSON缓存，并为新职位行应用高亮。
        Excel只保存指定字段，JSON保留全部字段。"""
        if not data_frames:
            logging.info("没有数据需要保存。")
            return
            
        try:
            # 为Excel创建简化的数据框
            excel_data_frames = {}
            for sheet_name, df in data_frames.items():
                # 创建Excel专用的数据框，只包含指定字段
                excel_df = df.copy()
                
                # 拆分description字段为团队介绍和岗位细节
                excel_df['团队介绍'] = excel_df['description'].apply(self._extract_team_intro)
                excel_df['岗位细节'] = excel_df['description'].apply(self._extract_job_details)
                
                # 添加职位链接
                excel_df['职位链接'] = excel_df['job_id'].apply(
                    lambda x: f"https://jobs.bytedance.com/campus/position/{x}/detail" if pd.notna(x) else ""
                )
                
                # 选择Excel需要的字段
                excel_columns = [
                    'title', '团队介绍', '岗位细节', 'requirement', 
                    'publish_time', 'code', '职位链接', 'highlight_time'
                ]
                
                # 只保留存在的列
                available_columns = [col for col in excel_columns if col in excel_df.columns]
                excel_df = excel_df[available_columns]
                
                # 重命名列为中文
                column_mapping = {
                    'title': '职位标题',
                    'requirement': '职位要求', 
                    'publish_time': '发布时间',
                    'code': '职位编号',
                    'highlight_time': '更新时间'
                }
                excel_df = excel_df.rename(columns=column_mapping)
                
                excel_data_frames[sheet_name] = excel_df
            
            # 保存Excel文件
            with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
                for sheet_name, df in excel_data_frames.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 应用Excel高亮
            workbook = openpyxl.load_workbook(self.filename)
            highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            highlight_stats = {}
            for sheet_name, df in excel_data_frames.items():
                if sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    new_count = 0
                    
                    # 高亮新增职位（更新时间不为空的行）
                    for idx, row in df.iterrows():
                        if pd.notna(row.get('更新时间')):
                            excel_row = idx + 2  # Excel行号从1开始，加上表头
                            for col in range(1, len(df.columns) + 1):
                                worksheet.cell(row=excel_row, column=col).fill = highlight_fill
                            new_count += 1
                    
                    highlight_stats[sheet_name] = new_count
            
            workbook.save(self.filename)
            logging.info(f"💾 Excel数据已保存并高亮至: {self.filename}")
            
            # 记录高亮统计信息
            for sheet_name, count in highlight_stats.items():
                if count > 0:
                    logging.info(f"   📌 工作表 '{sheet_name}' 高亮了 {count} 个新职位")
            
            # 保存JSON缓存（保留全部字段）
            self._save_json_cache(data_frames)

        except Exception as e:
            logging.error(f"⚠️ 保存Excel文件时出错: {e}")
            # 即使Excel保存失败，也尝试保存JSON缓存
            try:
                self._save_json_cache(data_frames)
            except Exception as cache_error:
                logging.error(f"⚠️ 保存JSON缓存时出错: {cache_error}")

    @staticmethod
    def _send_notification(summary: List[Dict[str, Any]]) -> None:
        """根据操作系统发送桌面通知，增强了稳定性和错误排查能力。"""
        total_new = sum(info.get('new_count', 0) for info in summary)
        total_count = sum(info.get('total_count', 0) for info in summary)
        current_time = datetime.now().strftime("%H:%M")

        if total_new > 0:
            title = "🎉 发现新职位!"
            details = "\\n".join(
                f"   • {info['task_name']}: +{info['new_count']} 个"
                for info in summary if info.get('new_count', 0) > 0
            )
            message = f"发现 {total_new} 个新丝瓜！\\n总计: {total_count} 个丝瓜\\n时间: {current_time}\\n\\n详情:\\n{details}"
            buttons = '{"稍后查看", "立即查看"}'
            default_button = '"立即查看"'
            action_script = f'do shell script "open \\"{OUTPUT_FILENAME}\\""'
        else:
            title = "✅ 监控完成"
            message = f"本次未发现新丝瓜。\\n总计: {total_count} 个丝瓜\\n时间: {current_time}"
            buttons = '{"确定"}'
            default_button = '"确定"'
            action_script = ""

        if sys.platform == "darwin":
            script = f'''
            try
                set response to display dialog "{message}" with title "{title}" buttons {buttons} default button {default_button} with icon note
                if button returned of response is "立即查看" then
                    {action_script}
                end if
            on error errMsg number errNum
                -- 此处留空，以便在用户取消或对话框自动消失时静默处理，避免Python端报错
            end try
            '''
            try:
                result = subprocess.run(
                    ['osascript', '-e', script], 
                    capture_output=True, text=True, timeout=120, check=False
                )
                if result.returncode != 0:
                    logging.error("⚠️ 发送 macOS 通知失败! AppleScript 执行出错。")
                    logging.error(f"   - 返回码: {result.returncode}")
                    logging.error(f"   - 错误输出: {result.stderr.strip()}")
                    logging.info("   - 请参考文档中的“通知问题排查指南”进行检查。")
                else:
                    logging.info("🔔 已成功触发桌面通知脚本。")
            except subprocess.TimeoutExpired:
                logging.warning("⚠️ 发送 macOS 通知超时。用户可能未在2分钟内响应对话框。")
            except Exception as e:
                logging.error(f"⚠️ 调用通知子进程时发生未知错误: {e}")
        else:
            logging.info("--- 桌面通知 (非macOS) ---")
            logging.info(f"标题: {title}")
            logging.info(message.replace("\\n", "\n"))
            logging.info("-------------------------")

    async def run_async(self, silent_mode: bool = False):
        """执行一次完整的职位监控流程（异步版）。"""
        start_time = datetime.now()
        logging.info(f"--- 开始监控 {start_time.strftime('%Y-%m-%d %H:%M:%S')} ---")
        
        # 确保输出目录存在
        DOCUMENTS_PATH.mkdir(exist_ok=True)
        
        existing_hashes, existing_dataframes = self._load_existing_hashes()
        
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=self.headless)
            tasks_to_run = [self._run_single_task_async(task, browser) for task in self.tasks]
            await asyncio.gather(*tasks_to_run)
            await browser.close()

        results = self._process_results(existing_hashes, existing_dataframes)
        data_frames = results["data_frames"]
        summary = results["summary"]
        
        self._save_and_highlight(data_frames)
        
        total_new = sum(info.get('new_count', 0) for info in summary)
        if not silent_mode or total_new > 0:
            logging.info("--- 监控结果 ---")
            for info in summary:
                logging.info(f"  - {info['task_name']}: 发现 {info.get('new_count', 0)} 个新丝瓜，共 {info.get('total_count', 0)} 个。")
            logging.info(f"总计新增: {total_new} 个")
        
        self._send_notification(summary)
        
        end_time = datetime.now()
        logging.info(f"--- 监控结束, 耗时: {(end_time - start_time).total_seconds():.2f} 秒 ---")

# --- 3. 主程序入口 ---
if __name__ == "__main__":
    try:
        is_silent = "--auto" in sys.argv
        monitor = JobMonitor(tasks=TASK_CONFIGS, filename=OUTPUT_FILENAME, headless=True)
        asyncio.run(monitor.run_async(silent_mode=is_silent))
        
    except KeyboardInterrupt:
        logging.info("\n⚠️ 程序被用户中断。")
        sys.exit(0)
    except Exception as e:
        if "Executable doesn't exist" in str(e):
            logging.critical("❌ Playwright 浏览器依赖未安装! 请运行 'playwright install' 命令。")
        else:
            logging.critical(f"\n❌ 程序发生严重错误: {e}", exc_info=True)
        sys.exit(1)
